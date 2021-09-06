import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup as bs
from selenium.webdriver import ActionChains
from selenium.webdriver.support.select import Select
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from datetime import datetime


def get_webdriver():
    chrome_options = Options()
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--incognito")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)

    driver = webdriver.Chrome(options=chrome_options)
    return driver


def get_data(page, container_no):
    today = datetime.today().strftime("%Y%m%d")
    '''
    엑셀에 저장
	* export.xlsx 와 동일하게 생성
    '''
    wb = Workbook()
    ws = wb.active
    header = ['CTR NO', 'CY GATE IN','ON BOARD', 'PORT ETA', 'PORT ATA', 'RAIL ATA', 'CTR PICK UP', 'CTR RETURN', 'END']
    # header = []
    # for head in page.select('#HistoryData > div > table > thead > tr > th'):
    #     header.append(head.text)
    ws.append(header)
    dataList = [0] * 9 #CTR 번호를 1열, 
    dataList[0] = container_no
    # tbody 저장
    rows = page.select('#HistoryData > div > table > tbody > tr')
    for i in range(len(rows)-1,-1,-1):
        tempRow = rows[i].select('td')
        event = tempRow[2].text.strip()
        eventDate = tempRow[3].text.strip()
        if event == 'Gate In Full':
            dataList[1] = eventDate
        elif 'Sailed' in event and 'POL' in event:
            dataList[2] = eventDate
        elif 'Arrived' in event and 'POD' in event:
            if 'Estimated' in event:
                dataList[3] = eventDate
            else:
                dataList[4] = eventDate
        elif 'Rail' in event and 'Arrived' in event:
            dataList[5] = eventDate
        elif 'Gate Out' in event:
            dataList[6] = eventDate



    ws.append(dataList)
    headerColor = PatternFill(start_color='c0c0c0', end_color='c0c0c0', fill_type='solid')
    for col in range(1,10):
        ws.column_dimensions[get_column_letter(col)].width = 20
        ws.cell(1,col).fill = headerColor


    wb.save(f'{today}-WestWood.xlsx')




def controll_page(driver, container_no):
    target = driver.find_element_by_css_selector('#ByContainer > div > div > div > textarea')

    action = ActionChains(driver)
    action.move_to_element(target)

    action.move_to_element(target).perform()

    target.click()
    target.clear()
    target.send_keys(container_no)

    driver.find_element_by_css_selector('#ByContainer > div > div > div > button').click()
    time.sleep(5)
    # driver.execute_script(f"var a = document.querySelectorAll('td > a');for (var i=0; i<a.length;i++){{ if (a[i].innerText==='{container_no}') {{a[i].click();break;}}}}")


if __name__ == '__main__':
    '''
    '''
    container_no = '#'
    url = 'https://www.wsl.com/Tracking/Public'

    driver = get_webdriver()
    try:
        driver.get(url)
        controll_page(driver, container_no)  # 화면 컨트롤 작성

        page = bs(driver.page_source, 'lxml')

        get_data(page, container_no)  # 데이터 추출후 프린트

    finally:
        driver.quit()
        # pass

    import sys

    sys.exit(0)
